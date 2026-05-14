"""
成长印记 - 学生管理 API 路由

注意：路由顺序很重要！
- 静态路由（如 /groups, /import）必须在动态路由（如 /{student_id}）之前定义
- 否则 FastAPI 会先匹配动态路由导致错误
"""
import io
import csv
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.models import Student, Class as ClassModel, Grade, School, User, ParentStudent, StudyGroup, StudentGroup
from app.schemas import (
    StudentCreate, StudentResponse, StudentListResponse,
    StudyGroupCreate, StudyGroupResponse,
    MessageResponse
)
from app.core.security import decode_token
from pydantic import BaseModel


class ParentStudentResponse(BaseModel):
    """家长绑定的学生信息"""
    id: str
    name: str
    student_number: Optional[str] = None
    class_id: str
    class_name: str
    grade_name: Optional[str] = None
    school_name: Optional[str] = None
    relation_type: str = "家长"

router = APIRouter(prefix="/students", tags=["学生管理"])
security = HTTPBearer()


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
) -> User:
    """获取当前用户"""
    token = credentials.credentials
    payload = decode_token(token)
    
    if not payload or payload.get("type") != "access":
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="无效的访问令牌"
        )
    
    user_id = payload.get("sub")
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="用户不存在"
        )
    
    return user


# ==================== 学生列表和创建 ====================

@router.get("/my", response_model=List[ParentStudentResponse])
async def get_my_students(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """获取当前家长绑定的学生信息（家长端）"""
    from app.core.security import decode_token
    
    token = credentials.credentials
    payload = decode_token(token)
    
    if not payload or payload.get("type") != "access":
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="无效的访问令牌"
        )
    
    user_id = payload.get("sub")
    role = payload.get("role")
    
    if role != "parent":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有家长角色可以访问此接口"
        )
    
    # 查询家长绑定的学生
    result = await db.execute(
        select(ParentStudent)
        .where(ParentStudent.parent_id == user_id)
    )
    parent_students = result.scalars().all()
    
    if not parent_students:
        return []
    
    # 构建响应
    response = []
    for ps in parent_students:
        # 获取学生信息
        s_result = await db.execute(
            select(Student)
            .options(selectinload(Student.class_).selectinload(ClassModel.grade).selectinload(Grade.school))
            .where(Student.id == ps.student_id)
        )
        student = s_result.scalar_one_or_none()
        
        if student and student.class_:
            class_name = student.class_.name
            grade_name = student.class_.grade.name if student.class_.grade else None
            school_name = student.class_.grade.school.name if student.class_.grade and student.class_.grade.school else None
        else:
            class_name = "未知班级"
            grade_name = None
            school_name = None
        
        response.append(ParentStudentResponse(
            id=str(student.id) if student else "",
            name=student.name if student else "未知学生",
            student_number=student.student_number if student else None,
            class_id=str(student.class_id) if student else "",
            class_name=class_name,
            grade_name=grade_name,
            school_name=school_name,
            relation_type=ps.relation_type or "家长"
        ))
    
    return response


@router.get("", response_model=StudentListResponse)
async def list_students(
    class_id: Optional[str] = Query(None, description="班级ID"),
    grade_id: Optional[str] = Query(None, description="年级ID"),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取学生列表"""
    query = select(Student).options(selectinload(Student.class_))
    
    if class_id:
        query = query.where(Student.class_id == class_id)
    elif grade_id:
        # 通过班级关联年级
        subquery = select(ClassModel.id).where(ClassModel.grade_id == grade_id)
        query = query.where(Student.class_id.in_(subquery))
    
    if keyword:
        query = query.where(Student.name.ilike(f"%{keyword}%"))
    
    # 统计总数
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar()
    
    # 分页
    query = query.order_by(Student.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)
    
    result = await db.execute(query)
    students = result.scalars().all()
    
    return StudentListResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=[StudentResponse.model_validate(s) for s in students]
    )


@router.post("", response_model=StudentResponse)
async def create_student(
    student_data: StudentCreate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """创建学生"""
    # 检查班级是否存在
    result = await db.execute(select(ClassModel).where(ClassModel.id == student_data.class_id))
    cls = result.scalar_one_or_none()
    
    if not cls:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="班级不存在"
        )
    
    student = Student(
        class_id=student_data.class_id,
        name=student_data.name,
        gender=student_data.gender,
        student_number=student_data.student_number,
        birth_date=student_data.birth_date
    )
    
    db.add(student)
    await db.commit()
    await db.refresh(student)
    
    return StudentResponse.model_validate(student)


@router.post("/import", response_model=MessageResponse)
async def import_students(
    class_id: str = Form(..., description="班级ID"),
    file: UploadFile = File(..., description="Excel/CSV文件"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """批量导入学生，支持组名字段"""
    import openpyxl
    
    # 检查班级是否存在
    result = await db.execute(select(ClassModel).where(ClassModel.id == class_id))
    cls = result.scalar_one_or_none()
    
    if not cls:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="班级不存在"
        )
    
    # 读取文件内容
    content = await file.read()
    
    students_created = 0
    students_updated = 0
    group_created = 0
    errors = []
    
    # 缓存已创建的小组 {组名: StudyGroup}
    group_cache: dict = {}
    
    # 预加载该班级已有的学习小组
    existing_groups = await db.execute(
        select(StudyGroup).where(StudyGroup.class_id == class_id)
    )
    for g in existing_groups.scalars().all():
        group_cache[g.name] = g
    
    try:
        if file.filename and file.filename.endswith('.csv'):
            # CSV格式
            text_content = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
        elif file.filename and (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
            # Excel格式 - 尝试解析
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        rows.append(dict(zip(headers, [str(v).strip() if v else '' for v in row])))
            except Exception as excel_err:
                # 如果xlsx解析失败，可能用户上传的是HTML伪装的xls
                # 尝试按CSV/TSV解析
                try:
                    text_content = content.decode('utf-8-sig')
                    # 检测是否是HTML格式
                    if '<html' in text_content.lower() or '<table' in text_content.lower():
                        # 解析HTML表格
                        import re
                        # 提取所有表格行
                        tr_pattern = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL | re.IGNORECASE)
                        td_pattern = re.compile(r'<t[dh][^>]*>(.*?)</t[dh]>', re.DOTALL | re.IGNORECASE)
                        tag_pattern = re.compile(r'<[^>]+>')
                        
                        trs = tr_pattern.findall(text_content)
                        headers = []
                        rows = []
                        for i, tr in enumerate(trs):
                            cells = td_pattern.findall(tr)
                            values = [tag_pattern.sub('', c).strip() for c in cells]
                            if i == 0:
                                headers = values
                            elif any(values):
                                rows.append(dict(zip(headers, values)))
                    else:
                        # 普通文本，按制表符分隔
                        reader = csv.DictReader(io.StringIO(text_content), delimiter='\t')
                        rows = list(reader)
                except Exception:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Excel文件解析失败，请使用 .xlsx 或 .csv 格式: {str(excel_err)}"
                    )
        else:
            # 无扩展名或其他格式，尝试自动检测
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        rows.append(dict(zip(headers, [str(v).strip() if v else '' for v in row])))
            except Exception:
                text_content = content.decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text_content))
                rows = list(reader)
        
        for idx, row in enumerate(rows, start=2):
            try:
                name = row.get('姓名') or row.get('学生姓名') or row.get('name', '')
                student_number = row.get('学号') or row.get('学生学号') or row.get('student_number', '')
                gender = row.get('性别') or row.get('gender', '')
                parent_phone = row.get('家长手机号') or row.get('家长电话') or row.get('parent_phone', '')
                group_name = row.get('组名') or row.get('学习小组') or row.get('group_name', '')
                
                if not name:
                    errors.append(f"第{idx}行: 缺少姓名")
                    continue
                
                # 清理空字符串为None
                student_number = student_number if student_number and student_number.strip() else None
                parent_phone = parent_phone if parent_phone and parent_phone.strip() else None
                group_name = group_name if group_name and group_name.strip() else None
                
                # 性别处理
                gender_value = 'male'
                if gender:
                    if gender in ['女', 'F', 'female', 'Female']:
                        gender_value = 'female'
                    elif gender in ['男', 'M', 'male', 'Male']:
                        gender_value = 'male'
                
                # 检查学号是否已存在
                existing = None
                if student_number:
                    result = await db.execute(
                        select(Student).where(
                            Student.class_id == class_id,
                            Student.student_number == student_number
                        )
                    )
                    existing = result.scalar_one_or_none()
                
                if existing:
                    existing.name = name
                    existing.gender = gender_value
                    # 处理小组变更
                    if group_name:
                        group = group_cache.get(group_name)
                        if not group:
                            group = StudyGroup(class_id=class_id, name=group_name)
                            db.add(group)
                            await db.flush()
                            group_cache[group_name] = group
                            group_created += 1
                        existing.study_group_id = group.id
                    students_updated += 1
                else:
                    # 处理学习小组
                    study_group_id = None
                    if group_name:
                        group = group_cache.get(group_name)
                        if not group:
                            group = StudyGroup(class_id=class_id, name=group_name)
                            db.add(group)
                            await db.flush()
                            group_cache[group_name] = group
                            group_created += 1
                        study_group_id = group.id
                    
                    student = Student(
                        class_id=class_id,
                        name=name,
                        gender=gender_value,
                        student_number=student_number,
                        study_group_id=study_group_id
                    )
                    db.add(student)
                    students_created += 1
                    
                    # 处理家长手机号 - 创建家长用户并关联
                    if parent_phone:
                        await db.flush()  # 确保student有id
                        # 检查是否已有该手机号的家长
                        parent_result = await db.execute(
                            select(User).where(User.phone == parent_phone)
                        )
                        parent = parent_result.scalar_one_or_none()
                        if not parent:
                            # 创建家长用户
                            from app.core.security import hash_password
                            parent = User(
                                phone=parent_phone,
                                password_hash=hash_password(f"cz{parent_phone[-4:]}"),
                                name=f"{name}家长",
                                role='parent'
                            )
                            db.add(parent)
                            await db.flush()
                        # 创建家长-学生关联
                        ps_result = await db.execute(
                            select(ParentStudent).where(
                                ParentStudent.parent_id == parent.id,
                                ParentStudent.student_id == student.id
                            )
                        )
                        if not ps_result.scalar_one_or_none():
                            ps = ParentStudent(
                                parent_id=parent.id,
                                student_id=student.id
                            )
                            db.add(ps)
                    
                    # 同时添加到student_groups关联表
                    if study_group_id:
                        await db.flush()
                        sg = StudentGroup(
                            student_id=student.id,
                            group_id=study_group_id
                        )
                        db.add(sg)
                    
            except Exception as e:
                errors.append(f"第{idx}行: {str(e)}")
        
        await db.commit()
        
        msg = f"导入完成：新增 {students_created} 人，更新 {students_updated} 人"
        if group_created > 0:
            msg += f"，新建小组 {group_created} 个"
        if errors:
            msg += f"，错误 {len(errors)} 条"
        
        return MessageResponse(message=msg)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"文件解析失败: {str(e)}"
        )


# ==================== 学习小组管理（静态路由，必须在 /{student_id} 之前）====================

@router.get("/groups", response_model=List[StudyGroupResponse])
async def list_study_groups(
    class_id: str = Query(..., description="班级ID"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取班级的学习小组列表"""
    result = await db.execute(
        select(StudyGroup)
        .options(selectinload(StudyGroup.students))
        .where(StudyGroup.class_id == class_id)
        .order_by(StudyGroup.sort_order)
    )
    groups = result.scalars().all()
    
    return [
        StudyGroupResponse(
            id=g.id,
            name=g.name,
            class_id=g.class_id,
            sort_order=g.sort_order,
            students=[StudentResponse.model_validate(s) for s in g.students]
        )
        for g in groups
    ]


@router.post("/groups", response_model=StudyGroupResponse)
async def create_study_group(
    group_data: StudyGroupCreate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """创建学习小组"""
    # 检查班级是否存在
    result = await db.execute(select(ClassModel).where(ClassModel.id == group_data.class_id))
    cls = result.scalar_one_or_none()
    
    if not cls:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="班级不存在"
        )
    
    # 获取最大排序号
    max_order_result = await db.execute(
        select(func.max(StudyGroup.sort_order)).where(StudyGroup.class_id == group_data.class_id)
    )
    max_order = max_order_result.scalar() or 0
    
    group = StudyGroup(
        class_id=group_data.class_id,
        name=group_data.name,
        sort_order=max_order + 1
    )
    
    db.add(group)
    await db.commit()
    await db.refresh(group)
    
    return StudyGroupResponse(
        id=group.id,
        name=group.name,
        class_id=group.class_id,
        sort_order=group.sort_order,
        students=[]
    )


@router.post("/groups/{group_id}/members", response_model=MessageResponse)
async def add_student_to_group(
    group_id: str,
    student_id: str = Form(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """将学生添加到学习小组"""
    # 检查小组是否存在
    result = await db.execute(select(StudyGroup).where(StudyGroup.id == group_id))
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学习小组不存在"
        )
    
    # 检查学生是否存在
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学生不存在"
        )
    
    # 更新学生的小组
    student.study_group_id = group_id
    await db.commit()
    
    return MessageResponse(message="添加成功")


@router.delete("/groups/{group_id}/members/{student_id}", response_model=MessageResponse)
async def remove_student_from_group(
    group_id: str,
    student_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """将学生从学习小组移除"""
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学生不存在"
        )
    
    student.study_group_id = None
    await db.commit()
    
    return MessageResponse(message="移除成功")


@router.delete("/groups/{group_id}", response_model=MessageResponse)
async def delete_study_group(
    group_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """删除学习小组"""
    result = await db.execute(select(StudyGroup).where(StudyGroup.id == group_id))
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学习小组不存在"
        )
    
    # 先将组内学生移出
    await db.execute(
        Student.__table__.update()
        .where(Student.study_group_id == group_id)
        .values(study_group_id=None)
    )
    
    await db.delete(group)
    await db.commit()
    
    return MessageResponse(message="删除成功")


# ==================== 学生详情、编辑、转班、删除（动态路由，必须在静态路由之后）====================

@router.get("/{student_id}", response_model=StudentResponse)
async def get_student(
    student_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取学生详情"""
    result = await db.execute(
        select(Student)
        .options(selectinload(Student.class_), selectinload(Student.study_group))
        .where(Student.id == student_id)
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "学生不存在")
    return StudentResponse.model_validate(student)


class StudentUpdate(BaseModel):
    name: Optional[str] = None
    student_number: Optional[str] = None
    study_group_id: Optional[str] = None
    status: Optional[str] = None  # active, leave, transfer, leave_school


@router.put("/{student_id}", response_model=StudentResponse)
async def update_student(
    student_id: str,
    data: StudentUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """编辑学生信息"""
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "学生不存在")
    
    update_data = data.model_dump(exclude_unset=True)
    
    # 学号唯一性检查
    if data.student_number is not None and data.student_number != student.student_number:
        existing = await db.execute(
            select(Student).where(
                Student.class_id == student.class_id,
                Student.student_number == data.student_number
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(400, "该学号在同班已存在")
    
    for key, value in update_data.items():
        if key == "study_group_id" and value:
            g = await db.execute(select(StudyGroup).where(StudyGroup.id == value))
            group = g.scalar_one_or_none()
            if not group or str(group.class_id) != str(student.class_id):
                raise HTTPException(400, "无效的学习小组")
        setattr(student, key, value)
    
    await db.commit()
    await db.refresh(student)
    return StudentResponse.model_validate(student)


class TransferRequest(BaseModel):
    target_class_id: str


@router.put("/{student_id}/transfer", response_model=MessageResponse)
async def transfer_student(
    student_id: str,
    data: TransferRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """学生转班"""
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "学生不存在")
    
    target = await db.execute(select(ClassModel).where(ClassModel.id == data.target_class_id))
    if not target.scalar_one_or_none():
        raise HTTPException(404, "目标班级不存在")
    
    # 清除原学习小组，转移班级
    student.study_group_id = None
    student.class_id = data.target_class_id
    student.student_number = None  # 转班后需重新设置学号
    
    # 删除原班student_groups关联
    await db.execute(
        StudentGroup.__table__.delete().where(StudentGroup.student_id == student_id)
    )
    
    await db.commit()
    return MessageResponse(message="转班成功，请重新设置学号")


@router.delete("/{student_id}", response_model=MessageResponse)
async def delete_student(
    student_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """删除学生及关联数据"""
    from app.models import TaskSubmission, SubmissionImage, TeacherAnnotation, EvaluationRecord, Achievement, AccessLog
    
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(404, "学生不存在")
    
    # 1. 查找提交记录
    sub_result = await db.execute(
        select(TaskSubmission.id).where(TaskSubmission.student_id == student_id)
    )
    submission_ids = [row[0] for row in sub_result.all()]
    if submission_ids:
        img_result = await db.execute(
            select(SubmissionImage.id).where(SubmissionImage.submission_id.in_(submission_ids))
        )
        image_ids = [row[0] for row in img_result.all()]
        if image_ids:
            await db.execute(
                TeacherAnnotation.__table__.delete().where(TeacherAnnotation.image_id.in_(image_ids))
            )
            await db.execute(
                SubmissionImage.__table__.delete().where(SubmissionImage.id.in_(image_ids))
            )
        await db.execute(
            TaskSubmission.__table__.delete().where(TaskSubmission.id.in_(submission_ids))
        )
    
    # 2. 删除评价记录
    await db.execute(
        EvaluationRecord.__table__.delete().where(EvaluationRecord.student_id == student_id)
    )
    
    # 3. 删除成就
    await db.execute(
        Achievement.__table__.delete().where(Achievement.student_id == student_id)
    )
    
    # 4. 删除访问日志
    await db.execute(
        AccessLog.__table__.delete().where(AccessLog.target_student_id == student_id)
    )
    
    # 5. 删除家长关联
    await db.execute(
        ParentStudent.__table__.delete().where(ParentStudent.student_id == student_id)
    )
    
    # 6. 删除学生-小组关联
    await db.execute(
        StudentGroup.__table__.delete().where(StudentGroup.student_id == student_id)
    )
    
    # 7. 删除学生
    await db.delete(student)
    await db.commit()
    
    return MessageResponse(message="学生及关联数据已删除")
